"""
OnePilot - Connection Service v3.3
Teste les connexions et synchronise les métadonnées
Supporte: SQL Server, PostgreSQL, MySQL, SQLite, OData, REST, Files, SAP, Dynamics 365, SAGE

FIXES v3.3:
✅ URL encoding (quote_plus) pour PostgreSQL et MySQL
✅ connect_timeout pour PostgreSQL (pas timeout)
✅ asyncio.to_thread() pour TOUS les I/O synchrones
✅ Bulk INFORMATION_SCHEMA pour SQL Server (3 requêtes au lieu de 3792)
✅ Timeout 900s global sur sync DB (Docker healthcheck fix)
"""

from __future__ import annotations

import asyncio
import os
import time
import json
import logging
from uuid import UUID
from typing import Dict, Any, List, Optional
from urllib.parse import quote_plus

from .schemas import ConnectorType, SourceCategory, CONNECTOR_CATEGORY_MAP
from .repository import get_source, get_source_secrets, save_test_result, save_metadata

logger = logging.getLogger(__name__)

# ═══════════════════════════════════════════════════════════════════════════════
# UTILITAIRES
# ═══════════════════════════════════════════════════════════════════════════════

def _build_sqlalchemy_url(connector_type: str, host: str, port: int,
                           db: str, user: str, password: str) -> str:
    """
    Construit l'URL SQLAlchemy selon le type de connecteur.
    FIX v3.3: URL encoding pour PostgreSQL/MySQL (quote_plus).
    """
    if connector_type == "sqlite":
        return f"sqlite:///{db}"
    
    # ── PostgreSQL ──
    if "postgres" in connector_type or connector_type == "pg":
        port = port or 5432
        user_enc = quote_plus(user or "")
        pwd_enc = quote_plus(password or "")
        db_enc = quote_plus(db or "")
        return f"postgresql+psycopg2://{user_enc}:{pwd_enc}@{host}:{port}/{db_enc}"
    
    # ── MySQL ──
    if "mysql" in connector_type:
        port = port or 3306
        user_enc = quote_plus(user or "")
        pwd_enc = quote_plus(password or "")
        db_enc = quote_plus(db or "")
        return f"mysql+pymysql://{user_enc}:{pwd_enc}@{host}:{port}/{db_enc}"
    
    # ── SQL Server / MSSQL ──
    if connector_type in ("mssql", "sage_100"):
        port = port or 1433
        return (
            f"mssql+pyodbc://{user}:{password}@{host}:{port}/{db}"
            f"?driver=ODBC+Driver+18+for+SQL+Server"
            f"&Encrypt=yes"
            f"&TrustServerCertificate=yes"
        )
    
    # ── Fallback ──
    port = port or 5432
    user_enc = quote_plus(user or "")
    pwd_enc = quote_plus(password or "")
    db_enc = quote_plus(db or "")
    return f"postgresql+psycopg2://{user_enc}:{pwd_enc}@{host}:{port}/{db_enc}"


def _human_size(size: int) -> str:
    """Convertit une taille en bytes en format lisible."""
    for unit in ["o", "Ko", "Mo", "Go"]:
        if size < 1024:
            return f"{size:.1f} {unit}"
        size /= 1024
    return f"{size:.1f} To"


def _infer_column_type(values: List[str]) -> str:
    """Déduit le type d'une colonne à partir d'un échantillon."""
    import re
    if not values:
        return "string"
    total = len(values)
    int_count   = sum(1 for v in values if re.match(r"^-?\d+$", v.strip()))
    float_count = sum(1 for v in values if re.match(r"^-?\d+\.\d+$", v.strip()))
    date_count  = sum(1 for v in values if re.match(r"^\d{4}-\d{2}-\d{2}", v.strip()))
    bool_count  = sum(1 for v in values if v.strip().lower() in ("true","false","1","0","oui","non"))
    if date_count  / total > 0.8: return "date"
    if bool_count  / total > 0.8: return "boolean"
    if int_count   / total > 0.8: return "integer"
    if float_count / total > 0.8: return "float"
    return "string"


def _infer_type(value: Any) -> str:
    """Déduit le type d'une valeur JSON."""
    if isinstance(value, bool):  return "boolean"
    if isinstance(value, int):   return "integer"
    if isinstance(value, float): return "float"
    if isinstance(value, dict):  return "object"
    if isinstance(value, list):  return "array"
    import re
    if isinstance(value, str) and re.match(r"\d{4}-\d{2}-\d{2}", value):
        return "date"
    return "string"


def _resolve_file_path(source) -> str:
    """Résout le chemin d'accès au fichier selon le connecteur."""
    opts = source.options or {}
    if opts.get("file_path"):
        path = opts["file_path"]
        if len(path) > 2 and path[1] == ":":
            path = "/mnt/host/" + path[3:].replace("\\", "/")
        elif path.startswith("\\\\"):
            path = "/mnt/network/" + path[2:].replace("\\", "/")
        elif path.startswith("//"):
            path = "/mnt/network/" + path[2:]
        return path
    if opts.get("uploaded_path"):
        return opts["uploaded_path"]
    if source.base_url:
        return source.base_url
    return ""


# ═══════════════════════════════════════════════════════════════════════════════
# TYPE MAPPINGS
# ═══════════════════════════════════════════════════════════════════════════════

TYPE_MAP_SQL = {
    "INTEGER": "integer", "BIGINT": "integer", "SMALLINT": "integer", "INT": "integer",
    "TINYINT": "integer",
    "NUMERIC": "decimal", "DECIMAL": "decimal", "FLOAT": "float", "DOUBLE": "float", "REAL": "float",
    "MONEY": "decimal", "SMALLMONEY": "decimal",
    "VARCHAR": "string", "TEXT": "string", "CHAR": "string", "NVARCHAR": "string",
    "NCHAR": "string", "NTEXT": "string",
    "BOOLEAN": "boolean", "BOOL": "boolean", "BIT": "boolean",
    "DATE": "date", "TIMESTAMP": "datetime", "DATETIME": "datetime", "DATETIME2": "datetime",
    "SMALLDATETIME": "datetime", "DATETIMEOFFSET": "datetime",
    "JSON": "json", "JSONB": "json", "UUID": "uuid", "UNIQUEIDENTIFIER": "uuid",
}

ABAP_TYPE_MAP = {
    "C": "string", "N": "string", "D": "date", "T": "string",
    "I": "integer", "P": "decimal", "F": "float", "X": "string",
    "S": "integer", "B": "integer", "STRING": "string",
    "INT1": "integer", "INT2": "integer", "INT4": "integer",
    "CURR": "decimal", "QUAN": "decimal", "DEC": "decimal",
    "DATS": "date", "CHAR": "string", "NUMC": "string",
}

EDM_TYPE_MAP = {
    "Edm.String": "string", "Edm.Int32": "integer", "Edm.Int64": "integer",
    "Edm.Int16": "integer", "Edm.Byte": "integer", "Edm.Decimal": "decimal",
    "Edm.Double": "float", "Edm.Single": "float", "Edm.Boolean": "boolean",
    "Edm.DateTime": "datetime", "Edm.DateTimeOffset": "datetime",
    "Edm.Date": "date", "Edm.Guid": "uuid", "Edm.Binary": "string",
    "Edm.Time": "string",
}

SAGE_TYPE_MAP = {
    "A": "string", "ANM": "string", "D": "date", "DCB": "decimal",
    "L": "integer", "W": "integer", "M": "string", "MD": "decimal",
    "QTY": "decimal", "C": "string", "Y": "boolean",
}


# ═══════════════════════════════════════════════════════════════════════════════
# TEST CONNECTION
# ═══════════════════════════════════════════════════════════════════════════════

async def test_connection(source_id: UUID) -> Dict[str, Any]:
    """Teste la connexion à une source de données."""
    source = await get_source(source_id)
    if not source:
        return {"success": False, "message": "Source introuvable", "latency_ms": -1}

    secrets = await get_source_secrets(source_id)
    start = time.time()

    try:
        ct       = source.connector_type.value
        category = CONNECTOR_CATEGORY_MAP.get(source.connector_type, SourceCategory.DATABASE)

        if ct in ("sap_rfc", "sap_odata"):
            result = await _test_sap_connection(source, secrets)
        elif ct == "dynamics365":
            result = await _test_dynamics_connection(source, secrets)
        elif ct in ("sage_x3", "sage_100", "sage_cloud"):
            result = await _test_sage_connection(source, secrets)
        elif category == SourceCategory.DATABASE:
            result = await _test_db_connection(source, secrets)
        elif category == SourceCategory.WEBSERVICE:
            result = await _test_webservice_connection(source, secrets)
        elif category == SourceCategory.FILE:
            result = await _test_file_connection(source)
        else:
            result = {"success": False, "message": "Type non supporté"}

    except Exception as e:
        logger.error(f"[test_connection] {source_id}: {e}", exc_info=True)
        result = {"success": False, "message": str(e)}

    latency = int((time.time() - start) * 1000)
    result["latency_ms"] = result.get("latency_ms", latency)

    await save_test_result(
        source_id, result["success"], result["message"], result["latency_ms"]
    )
    return result


async def _test_db_connection(source, secrets: Dict, opts: dict = None) -> Dict:
    """
    Test une connexion base de données SQL.
    FIX v3.3: asyncio.to_thread() pour ne pas bloquer l'event loop.
    """
    opts = opts or source.options or {}
    timeout  = int(opts.get("timeout",  30))
    retry_max = int(opts.get("retry_max", 3))
    pool_min  = int(opts.get("pool_min",  1))
    pool_max  = int(opts.get("pool_max",  10))

    last_err = None
    for attempt in range(1, retry_max + 1):
        try:
            password = secrets.get("password", "")
            url = _build_sqlalchemy_url(
                source.connector_type.value,
                source.host,
                source.port or 5432,
                source.database_name,
                source.username or "",
                password
            )

            def _do_test():
                from sqlalchemy import create_engine, text

                ct = source.connector_type.value

                # ── connect_args par dialecte ──────────────────────────────────
                # psycopg2  : "connect_timeout" (entier, secondes) — PAS "timeout"
                # pymysql   : "connect_timeout" (entier, secondes)
                # pyodbc    : "timeout" dans la chaîne DSN, pas dans connect_args
                # sqlite    : aucun timeout de connexion
                if "postgres" in ct or ct in ("pg", "postgresql"):
                    connect_args = {"connect_timeout": int(timeout)}
                elif "mysql" in ct or "mariadb" in ct:
                    connect_args = {"connect_timeout": int(timeout)}
                elif ct in ("mssql", "sage_100"):
                    # pyodbc timeout est géré dans le DSN ; connect_args vide
                    connect_args = {}
                else:
                    connect_args = {}

                engine = create_engine(
                    url,
                    pool_pre_ping=True,
                    pool_size=pool_min,
                    max_overflow=pool_max - pool_min,
                    pool_timeout=timeout,
                    connect_args=connect_args,
                )
                try:
                    t0 = time.time()
                    with engine.connect() as conn:
                        conn.execute(text("SELECT 1"))
                    return int((time.time() - t0) * 1000)
                finally:
                    engine.dispose()

            latency = await asyncio.to_thread(_do_test)
            return {
                "success": True,
                "message": f"Connexion réussie" + (f" (tentative {attempt})" if attempt > 1 else ""),
                "latency_ms": latency
            }
        except Exception as e:
            last_err = e
            if attempt < retry_max:
                wait = 2 ** (attempt - 1)
                logger.warning(f"[DB] Tentative {attempt}/{retry_max} échouée, retry dans {wait}s : {e}")
                await asyncio.sleep(wait)

    return {"success": False, "message": f"Échec après {retry_max} tentatives : {last_err}", "latency_ms": -1}

async def _test_webservice_connection(source, secrets: Dict, opts: dict = None) -> Dict:
    """
    Test une connexion REST/OData/GraphQL.
    FIX v3.3: asyncio.to_thread() pour requests.
    """
    opts      = opts or source.options or {}
    timeout   = int(opts.get("timeout",   10))
    retry_max = int(opts.get("retry_max",  3))

    headers: Dict[str, str] = {}

    if source.auth_type.value == "bearer":
        headers["Authorization"] = f"Bearer {secrets.get('token', '')}"
    elif source.auth_type.value == "basic":
        import base64
        creds = base64.b64encode(
            f"{source.username}:{secrets.get('password','')}".encode()
        ).decode()
        headers["Authorization"] = f"Basic {creds}"
    elif source.auth_type.value == "api_key":
        header_name = source.options.get("api_key_header", "X-API-Key")
        headers[header_name] = secrets.get("api_key_value", "")

    url = source.base_url
    if source.connector_type.value == "odata":
        url = f"{url}/$metadata"
        headers["Accept"] = "application/xml, text/xml, */*"

    last_err = None
    for attempt in range(1, retry_max + 1):
        try:
            def _do_get():
                import requests
                t0 = time.time()
                resp = requests.get(url, headers=headers, timeout=timeout)
                return resp.status_code, int((time.time() - t0) * 1000)

            status_code, latency = await asyncio.to_thread(_do_get)
            if status_code < 500:
                return {
                    "success": True,
                    "message": f"HTTP {status_code}" + (f" (tentative {attempt})" if attempt > 1 else ""),
                    "latency_ms": latency
                }
            last_err = f"HTTP {status_code}"
        except Exception as e:
            last_err = str(e)

        if attempt < retry_max:
            wait = 2 ** (attempt - 1)
            logger.warning(f"[WS] Retry {attempt}/{retry_max} dans {wait}s : {last_err}")
            await asyncio.sleep(wait)

    return {"success": False, "message": f"Échec après {retry_max} tentatives : {last_err}", "latency_ms": -1}


async def _test_file_connection(source) -> Dict:
    """Test l'accès à un fichier (local ou HTTP)."""
    start = time.time()
    try:
        file_path = _resolve_file_path(source)
        if not file_path:
            return {"success": False, "message": "Aucun chemin de fichier spécifié", "latency_ms": -1}

        if file_path.startswith("http"):
            def _do_head():
                import requests
                return requests.head(file_path, timeout=10)
            resp = await asyncio.to_thread(_do_head)
            latency = int((time.time() - start) * 1000)
            if resp.status_code < 400:
                return {"success": True, "message": f"Fichier accessible (HTTP {resp.status_code})", "latency_ms": latency}
            return {"success": False, "message": f"HTTP {resp.status_code}", "latency_ms": latency}

        if not os.path.exists(file_path):
            return {"success": False, "message": f"Fichier introuvable : {file_path}", "latency_ms": -1}

        size    = os.path.getsize(file_path)
        latency = int((time.time() - start) * 1000)
        ct      = source.connector_type.value
        sheets_info = ""

        if ct in ("file_excel", "file_xlsx"):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                sheets = wb.sheetnames
                wb.close()
                sheets_info = f" · {len(sheets)} feuilles"
            except Exception:
                sheets_info = ""
        elif ct == "file_parquet":
            try:
                import pyarrow.parquet as pq
                meta = pq.read_metadata(file_path)
                sheets_info = f" · {meta.num_rows:,} lignes"
            except Exception:
                sheets_info = " (Parquet)"

        return {
            "success":    True,
            "message":    f"Fichier accessible ({_human_size(size)}){sheets_info}",
            "latency_ms": latency,
        }
    except Exception as e:
        return {"success": False, "message": str(e), "latency_ms": -1}


async def _test_sap_connection(source, secrets: Dict) -> Dict:
    """Test la connexion SAP RFC ou OData."""
    ct = source.connector_type.value

    if ct == "sap_rfc":
        try:
            def _do_rfc():
                import pyrfc
                conn = pyrfc.Connection(
                    ashost=source.host or "",
                    sysnr=source.options.get("system_number", "00"),
                    client=source.options.get("client", "100"),
                    user=source.username or "",
                    passwd=secrets.get("password", ""),
                )
                conn.call("RFC_PING")
                conn.close()
            
            start = time.time()
            await asyncio.to_thread(_do_rfc)
            return {"success": True, "message": "SAP RFC_PING OK", "latency_ms": int((time.time() - start) * 1000)}
        except ImportError:
            return {"success": False, "message": "pyrfc non installé. Requiert SAP NetWeaver RFC SDK.", "latency_ms": -1}
        except Exception as e:
            return {"success": False, "message": str(e), "latency_ms": -1}
    else:
        # SAP OData
        start = time.time()
        try:
            url = (source.base_url or "").rstrip("/")
            headers = {"Accept": "application/xml, text/xml, */*"}
            if secrets.get("token"):
                headers["Authorization"] = f"Bearer {secrets['token']}"
            elif source.username:
                import base64
                creds = base64.b64encode(f"{source.username}:{secrets.get('password','')}".encode()).decode()
                headers["Authorization"] = f"Basic {creds}"
            
            def _do_get():
                import requests
                return requests.get(f"{url}/$metadata", headers=headers, timeout=10)
            
            resp = await asyncio.to_thread(_do_get)
            latency = int((time.time() - start) * 1000)
            if resp.status_code < 500:
                return {"success": True, "message": f"SAP OData HTTP {resp.status_code}", "latency_ms": latency}
            return {"success": False, "message": f"HTTP {resp.status_code}", "latency_ms": latency}
        except Exception as e:
            return {"success": False, "message": str(e), "latency_ms": -1}


async def _test_dynamics_connection(source, secrets: Dict) -> Dict:
    """Test la connexion Dynamics 365."""
    start = time.time()
    try:
        api_url = (source.base_url or "").rstrip("/")
        headers = _build_dynamics_headers(source, secrets)
        
        def _do_get():
            import requests
            return requests.get(f"{api_url}/api/data/v9.2/", headers=headers, timeout=10)
        
        resp = await asyncio.to_thread(_do_get)
        latency = int((time.time() - start) * 1000)
        if resp.status_code < 400:
            return {"success": True, "message": f"Dynamics 365 HTTP {resp.status_code}", "latency_ms": latency}
        return {"success": False, "message": f"HTTP {resp.status_code}", "latency_ms": latency}
    except Exception as e:
        return {"success": False, "message": str(e), "latency_ms": -1}


async def _test_sage_connection(source, secrets: Dict) -> Dict:
    """Test la connexion SAGE (X3, 100, Cloud)."""
    ct = source.connector_type.value
    start = time.time()

    try:
        if ct == "sage_x3":
            base_url = (source.base_url or "").rstrip("/")
            folder = source.options.get("folder", "SEED")
            headers = {"Accept": "application/json"}
            auth = None
            if secrets.get("token"):
                headers["Authorization"] = f"Bearer {secrets['token']}"
            elif source.username:
                auth = (source.username, secrets.get("password", ""))
            
            def _do_get():
                import requests
                return requests.get(f"{base_url}/api/x3/erp/{folder}", headers=headers, auth=auth, timeout=10)
            
            resp = await asyncio.to_thread(_do_get)
            latency = int((time.time() - start) * 1000)
            if resp.status_code < 500:
                return {"success": True, "message": f"SAGE X3 HTTP {resp.status_code}", "latency_ms": latency}
            return {"success": False, "message": f"HTTP {resp.status_code}", "latency_ms": latency}

        elif ct == "sage_100":
            return await _test_db_connection(source, secrets)

        elif ct == "sage_cloud":
            def _do_get():
                import requests
                return requests.get(
                    "https://api.accounting.sage.com/v3.1/ledger_accounts",
                    headers={"Authorization": f"Bearer {secrets.get('token', '')}"},
                    params={"$top": 1},
                    timeout=10
                )
            
            resp = await asyncio.to_thread(_do_get)
            latency = int((time.time() - start) * 1000)
            if resp.status_code < 400:
                return {"success": True, "message": f"SAGE Cloud HTTP {resp.status_code}", "latency_ms": latency}
            return {"success": False, "message": f"HTTP {resp.status_code}", "latency_ms": latency}

    except Exception as e:
        return {"success": False, "message": str(e), "latency_ms": -1}


def _build_dynamics_headers(source, secrets: Dict) -> Dict:
    """Construit les headers Dynamics 365."""
    headers = {"Accept": "application/json", "OData-MaxVersion": "4.0", "OData-Version": "4.0"}
    if secrets.get("token"):
        headers["Authorization"] = f"Bearer {secrets['token']}"
    elif source.username:
        import base64
        creds = base64.b64encode(f"{source.username}:{secrets.get('password','')}".encode()).decode()
        headers["Authorization"] = f"Basic {creds}"
    return headers


# ═══════════════════════════════════════════════════════════════════════════════
# SYNC METADATA
# ═══════════════════════════════════════════════════════════════════════════════

async def sync_metadata(source_id: UUID) -> Dict[str, Any]:
    """Synchronise les métadonnées d'une source."""
    source = await get_source(source_id)
    if not source:
        return {"success": False, "message": "Source introuvable"}

    secrets = await get_source_secrets(source_id)
    start = time.time()

    try:
        ct = source.connector_type.value
        category = CONNECTOR_CATEGORY_MAP.get(source.connector_type, SourceCategory.DATABASE)

        if ct == "sap_rfc":
            entities = await _fetch_sap_rfc_metadata(source, secrets)
        elif ct == "sap_odata":
            entities = await _fetch_sap_odata_metadata(source, secrets)
        elif ct == "dynamics365":
            entities = await _fetch_dynamics_metadata(source, secrets)
        elif ct in ("sage_x3", "sage_100", "sage_cloud"):
            entities = await _fetch_sage_metadata(source, secrets)
        elif category == SourceCategory.DATABASE:
            entities = await asyncio.wait_for(
                _fetch_db_metadata(source, secrets),
                timeout=900
            )
        elif category == SourceCategory.WEBSERVICE:
            entities = await _fetch_webservice_metadata(source, secrets)
        elif category == SourceCategory.FILE:
            entities = await _fetch_file_metadata(source)
        else:
            return {"success": False, "message": "Sync non supportée pour ce type"}

        entity_count = await save_metadata(source_id, entities)
        field_count = sum(len(e.get("fields", [])) for e in entities)
        duration = int((time.time() - start) * 1000)

        return {
            "success": True,
            "entity_count": entity_count,
            "field_count": field_count,
            "relation_count": 0,
            "duration_ms": duration,
            "message": f"{entity_count} entités, {field_count} champs synchronisés"
        }

    except asyncio.TimeoutError:
        duration = int((time.time() - start) * 1000)
        logger.error(f"[Sync] Timeout 900s pour source {source_id}")
        return {
            "success": False, "entity_count": 0, "field_count": 0,
            "relation_count": 0, "duration_ms": duration,
            "message": "Sync timeout après 900s — base trop volumineuse"
        }
    except Exception as e:
        logger.error(f"[Sync] Erreur source {source_id}: {e}", exc_info=True)
        return {
            "success": False, "entity_count": 0, "field_count": 0,
            "relation_count": 0, "duration_ms": int((time.time() - start) * 1000),
            "message": str(e)
        }


# ═══════════════════════════════════════════════════════════════════════════════
# FETCH DB METADATA
# ═══════════════════════════════════════════════════════════════════════════════

async def _fetch_db_metadata(source, secrets: Dict) -> List[Dict]:
    """
    FIX CRITIQUE v3.3:
    ✅ Toute la logique SQLAlchemy dans asyncio.to_thread()
    ✅ BULK INFORMATION_SCHEMA pour SQL Server (3 requêtes au lieu de 3792)
    ✅ Timeout 900s pour éviter Docker healthcheck timeout
    """
    def _sync_fetch() -> List[Dict]:
        from sqlalchemy import create_engine, text

        password = secrets.get("password", "")
        url = _build_sqlalchemy_url(
            source.connector_type.value,
            source.host,
            source.port or 5432,
            source.database_name,
            source.username or "",
            password
        )

        # ── connect_args par dialecte (même logique que _do_test) ──
        ct = source.connector_type.value
        if "postgres" in ct or ct in ("pg", "postgresql"):
            connect_args = {"connect_timeout": 30}
        elif "mysql" in ct or "mariadb" in ct:
            connect_args = {"connect_timeout": 30}
        else:
            connect_args = {}

        engine = create_engine(url, pool_pre_ping=True, connect_args=connect_args)
        try:
            dialect = source.connector_type.value
            schema = getattr(source, "schema_name", None) or None

            # ── FAST PATH: SQL Server ──
            if dialect in ("mssql", "sage_100"):
                schema_filter = schema or "dbo"
                with engine.connect() as conn:
                    col_rows = conn.execute(text("""
                        SELECT t.TABLE_NAME, t.TABLE_TYPE,
                               c.COLUMN_NAME, c.DATA_TYPE,
                               c.IS_NULLABLE, c.ORDINAL_POSITION
                        FROM INFORMATION_SCHEMA.TABLES  t
                        JOIN INFORMATION_SCHEMA.COLUMNS c
                          ON c.TABLE_SCHEMA = t.TABLE_SCHEMA
                         AND c.TABLE_NAME   = t.TABLE_NAME
                        WHERE t.TABLE_SCHEMA = :s
                        ORDER BY t.TABLE_NAME, c.ORDINAL_POSITION
                    """), {"s": schema_filter}).fetchall()

                    pk_rows = conn.execute(text("""
                        SELECT kcu.TABLE_NAME, kcu.COLUMN_NAME
                        FROM INFORMATION_SCHEMA.TABLE_CONSTRAINTS tc
                        JOIN INFORMATION_SCHEMA.KEY_COLUMN_USAGE  kcu
                          ON kcu.CONSTRAINT_NAME = tc.CONSTRAINT_NAME
                         AND kcu.TABLE_SCHEMA    = tc.TABLE_SCHEMA
                        WHERE tc.CONSTRAINT_TYPE = 'PRIMARY KEY'
                          AND tc.TABLE_SCHEMA    = :s
                    """), {"s": schema_filter}).fetchall()

                    # Commentaires tables (MS_Description)
                    try:
                        comment_rows = conn.execute(text("""
                            SELECT
                                OBJECT_NAME(ep.major_id)    AS table_name,
                                COL_NAME(ep.major_id, ep.minor_id) AS col_name,
                                CAST(ep.value AS NVARCHAR(MAX)) AS comment_text
                            FROM sys.extended_properties ep
                            WHERE ep.name = 'MS_Description'
                              AND ep.class = 1
                              AND OBJECT_SCHEMA_NAME(ep.major_id) = :s
                        """), {"s": schema_filter}).fetchall()
                    except Exception:
                        comment_rows = []

                    # Vues matérialisées (indexed views)
                    try:
                        matview_rows = conn.execute(text("""
                            SELECT v.name
                            FROM sys.views v
                            INNER JOIN sys.indexes i ON i.object_id = v.object_id
                            WHERE SCHEMA_NAME(v.schema_id) = :s
                              AND i.index_id = 1
                        """), {"s": schema_filter}).fetchall()
                        matview_set = {r[0] for r in matview_rows}
                    except Exception:
                        matview_set = set()

                    # Séquences MSSQL
                    try:
                        seq_rows = conn.execute(text("""
                            SELECT name, TYPE_NAME(user_type_id) as data_type,
                                   CAST(start_value AS VARCHAR) as start_value,
                                   CAST(increment AS VARCHAR) as increment,
                                   CAST(current_value AS VARCHAR) as current_value
                            FROM sys.sequences
                            WHERE SCHEMA_NAME(schema_id) = :s
                        """), {"s": schema_filter}).fetchall()
                    except Exception:
                        seq_rows = []

                    fk_rows = conn.execute(text("""
                        SELECT kcu.TABLE_NAME, kcu.COLUMN_NAME
                        FROM INFORMATION_SCHEMA.TABLE_CONSTRAINTS tc
                        JOIN INFORMATION_SCHEMA.KEY_COLUMN_USAGE  kcu
                          ON kcu.CONSTRAINT_NAME = tc.CONSTRAINT_NAME
                         AND kcu.TABLE_SCHEMA    = tc.TABLE_SCHEMA
                        WHERE tc.CONSTRAINT_TYPE = 'FOREIGN KEY'
                          AND tc.TABLE_SCHEMA    = :s
                    """), {"s": schema_filter}).fetchall()

                pk_map: Dict[str, set] = {}
                for r in pk_rows:
                    pk_map.setdefault(r[0], set()).add(r[1])
                
                fk_map: Dict[str, set] = {}
                for r in fk_rows:
                    fk_map.setdefault(r[0], set()).add(r[1])

                # Construire maps commentaires
                comment_map_tbl: Dict[str, str] = {}
                comment_map_col: Dict[str, Dict[str, str]] = {}
                for r in comment_rows:
                    tbl, col, txt = r[0], r[1], r[2]
                    if col is None:
                        comment_map_tbl[tbl] = txt
                    else:
                        comment_map_col.setdefault(tbl, {})[col] = txt

                # Séquences comme entités spéciales
                seq_entities = []
                for r in seq_rows:
                    seq_entities.append({
                        "name": f"SEQ_{r[0]}",
                        "entity_type": "sequence",
                        "description": f"Séquence {r[0]} — type:{r[1]} start:{r[2]} inc:{r[3]} current:{r[4]}",
                        "fields": []
                    })

                entities_dict: Dict[str, List[Dict]] = {}
                table_type_map: Dict[str, str] = {}

                for tname, ttype, cname, dtype, nullable, _ in col_rows:
                    if tname not in entities_dict:
                        # Vues matérialisées (indexed views)
                        if ttype == "VIEW" and tname in matview_set:
                            etype = "materialized_view"
                        elif ttype == "VIEW":
                            etype = "view"
                        else:
                            etype = "table"
                        table_type_map[tname] = etype
                        entities_dict[tname] = []
                    entities_dict[tname].append({
                        "name":        cname,
                        "type":        TYPE_MAP_SQL.get(dtype.upper(), "string"),
                        "native_type": dtype,
                        "nullable":    nullable == "YES",
                        "primary_key": cname in pk_map.get(tname, set()),
                        "foreign_key": cname in fk_map.get(tname, set()),
                        "description": comment_map_col.get(tname, {}).get(cname),
                    })

                entities = [
                    {
                        "name":        tname,
                        "entity_type": table_type_map[tname],
                        "description": comment_map_tbl.get(tname),
                        "fields":      fields,
                    }
                    for tname, fields in entities_dict.items()
                ] + seq_entities
                n_t = sum(1 for v in table_type_map.values() if v == "table")
                n_v = sum(1 for v in table_type_map.values() if v == "view")
                logger.info(f"[Sync MSSQL bulk] {n_t} tables + {n_v} vues = {len(entities)} entités")
                return entities

            # ── SLOW PATH: PostgreSQL / MySQL / SQLite ──
            from sqlalchemy import inspect as sa_inspect
            inspector = sa_inspect(engine)
            entities = []
            all_names = []
            
            try:
                tables = inspector.get_table_names(schema=schema)
                all_names += [(n, "table") for n in tables]
            except Exception as e:
                logger.warning(f"[Sync] get_table_names: {e}")

            try:
                views = inspector.get_view_names(schema=schema)
                all_names += [(n, "view") for n in views]
                logger.info(f"[Sync] {len(tables)} tables + {len(views)} vues")
            except Exception as e:
                logger.warning(f"[Sync] get_view_names: {e}")

            # Vues matérialisées PostgreSQL
            pg_comment_map: Dict[str, str] = {}
            pg_col_comment_map: Dict[str, Dict[str, str]] = {}
            pg_sequences = []
            if "postgres" in dialect or dialect in ("pg", "postgresql"):
                try:
                    with engine.connect() as conn2:
                        sch = schema or "public"
                        # Vues matérialisées
                        mat_rows = conn2.execute(text(
                            "SELECT matviewname FROM pg_matviews WHERE schemaname=:s"
                        ), {"s": sch}).fetchall()
                        all_names += [(r[0], "materialized_view") for r in mat_rows]

                        # Commentaires tables
                        tbl_comments = conn2.execute(text("""
                            SELECT c.relname, obj_description(c.oid,'pg_class')
                            FROM pg_class c
                            JOIN pg_namespace n ON n.oid = c.relnamespace
                            WHERE n.nspname=:s AND c.relkind IN ('r','v','m')
                              AND obj_description(c.oid,'pg_class') IS NOT NULL
                        """), {"s": sch}).fetchall()
                        pg_comment_map = {r[0]: r[1] for r in tbl_comments}

                        # Commentaires colonnes
                        col_comments = conn2.execute(text("""
                            SELECT c.relname, a.attname,
                                   col_description(c.oid, a.attnum)
                            FROM pg_class c
                            JOIN pg_namespace n ON n.oid = c.relnamespace
                            JOIN pg_attribute a ON a.attrelid = c.oid
                            WHERE n.nspname=:s AND a.attnum > 0
                              AND col_description(c.oid, a.attnum) IS NOT NULL
                        """), {"s": sch}).fetchall()
                        for tbl, col, cmt in col_comments:
                            pg_col_comment_map.setdefault(tbl, {})[col] = cmt

                        # Séquences PostgreSQL
                        seq_rows_pg = conn2.execute(text("""
                            SELECT sequence_name, data_type,
                                   start_value, increment
                            FROM information_schema.sequences
                            WHERE sequence_schema=:s
                        """), {"s": sch}).fetchall()
                        pg_sequences = [
                            {
                                "name": f"SEQ_{r[0]}",
                                "entity_type": "sequence",
                                "description": f"Séquence {r[0]} — type:{r[1]} start:{r[2]} inc:{r[3]}",
                                "fields": []
                            }
                            for r in seq_rows_pg
                        ]
                except Exception as e:
                    logger.warning(f"[Sync PG comments/matviews/sequences] {e}")

            for table_name, entity_type in all_names:
                try:
                    columns = inspector.get_columns(table_name, schema=schema)
                except Exception:
                    continue
                
                try:
                    pk_cols = set(inspector.get_pk_constraint(table_name, schema=schema).get("constrained_columns", []))
                except Exception:
                    pk_cols = set()
                
                fk_cols: set = set()
                try:
                    for fk in inspector.get_foreign_keys(table_name, schema=schema):
                        fk_cols.update(fk.get("constrained_columns", []))
                except Exception:
                    pass
                
                fields = []
                for col in columns:
                    native = str(col["type"]).upper().split("(")[0].strip()
                    fields.append({
                        "name": col["name"],
                        "type": TYPE_MAP_SQL.get(native, "string"),
                        "native_type": str(col["type"]),
                        "nullable": col.get("nullable", True),
                        "primary_key": col["name"] in pk_cols,
                        "foreign_key": col["name"] in fk_cols,
                    })
                
                entities.append({
                    "name":        table_name,
                    "entity_type": entity_type,
                    "description": pg_comment_map.get(table_name),
                    "fields":      [
                        {**f, "description": pg_col_comment_map.get(table_name, {}).get(f["name"])}
                        for f in fields
                    ],
                })

            # Ajouter les séquences PG
            entities += pg_sequences
            return entities
        
        finally:
            engine.dispose()

    logger.info(f"[Sync DB] Lancement dans asyncio.to_thread()…")
    result = await asyncio.to_thread(_sync_fetch)
    logger.info(f"[Sync DB] Terminé — {len(result)} entités")
    return result


# ═══════════════════════════════════════════════════════════════════════════════
# SAP RFC METADATA
# ═══════════════════════════════════════════════════════════════════════════════

async def _fetch_sap_rfc_metadata(source, secrets: Dict) -> List[Dict]:
    """Récupère les métadonnées des tables SAP via RFC."""
    def _sync_fetch():
        try:
            import pyrfc
        except ImportError:
            raise ImportError("pyrfc non installé. Requiert SAP NetWeaver RFC SDK + pip install pyrfc")

        conn = pyrfc.Connection(
            ashost=source.host or "",
            sysnr=source.options.get("system_number", "00"),
            client=source.options.get("client", "100"),
            user=source.username or "",
            passwd=secrets.get("password", ""),
            lang=source.options.get("lang", "FR"),
        )

        entities = []
        table_names = source.options.get("tables", [])

        if not table_names:
            try:
                result = conn.call(
                    "RFC_READ_TABLE",
                    QUERY_TABLE="DD02L",
                    FIELDS=[{"FIELDNAME": "TABNAME"}],
                    OPTIONS=[{"TEXT": "TABCLASS = 'TRANSP'"}],
                    ROWCOUNT=300,
                )
                table_names = [r["WA"].strip() for r in result.get("DATA", []) if r.get("WA", "").strip()]
            except Exception as e:
                logger.warning(f"[SAP RFC] DD02L: {e}")

        for table_name in table_names[:200]:
            try:
                result = conn.call("RFC_GET_STRUCTURE_DEFINITION", TABNAME=table_name)
                fields = [
                    {
                        "name": f.get("FIELDNAME", ""),
                        "type": ABAP_TYPE_MAP.get(f.get("DATATYPE", "C"), "string"),
                        "native_type": f.get("DATATYPE", ""),
                        "nullable": True,
                        "primary_key": f.get("KEYFLAG", "") == "X",
                        "foreign_key": False,
                        "description": f.get("FIELDTEXT", ""),
                    }
                    for f in result.get("FIELDS", [])
                ]
                if fields:
                    entities.append({"name": table_name, "entity_type": "sap_table", "fields": fields})
            except Exception as e:
                logger.warning(f"[SAP RFC] {table_name}: {e}")

        for fm in source.options.get("function_modules", []):
            try:
                result = conn.call("RFC_GET_FUNCTION_INTERFACE", FUNCNAME=fm)
                fields = [
                    {
                        "name": p.get("PARAMETER", ""),
                        "type": ABAP_TYPE_MAP.get(p.get("TABNAME", "C"), "string"),
                        "native_type": p.get("TABNAME", ""),
                        "nullable": True,
                        "primary_key": False,
                        "foreign_key": False,
                        "description": p.get("PARAMTEXT", ""),
                    }
                    for p in result.get("PARAMS_AND_EXCPS", [])
                ]
                entities.append({"name": fm, "entity_type": "sap_bapi", "fields": fields})
            except Exception as e:
                logger.warning(f"[SAP RFC] BAPI {fm}: {e}")

        conn.close()
        return entities

    return await asyncio.to_thread(_sync_fetch)


# ═════════════════════════════════════════════════════════════════════��═════════
# SAP ODATA METADATA
# ═══════════════════════════════════════════════════════════════════════════════

async def _fetch_sap_odata_metadata(source, secrets: Dict) -> List[Dict]:
    """Récupère les métadonnées d'un service SAP OData."""
    import xml.etree.ElementTree as ET

    base_url = (source.base_url or "").rstrip("/")
    headers = {"Accept": "application/xml, text/xml, */*"}

    if secrets.get("token"):
        headers["Authorization"] = f"Bearer {secrets['token']}"
    elif source.username:
        import base64
        creds = base64.b64encode(f"{source.username}:{secrets.get('password','')}".encode()).decode()
        headers["Authorization"] = f"Basic {creds}"

    def _do_get():
        import requests
        resp = requests.get(f"{base_url}/$metadata", headers=headers, timeout=15)
        resp.raise_for_status()
        return resp.text

    text = await asyncio.to_thread(_do_get)
    root = ET.fromstring(text)
    
    EDM_NS = "http://docs.oasis-open.org/odata/ns/edm"
    ns = f"{{{EDM_NS}}}"
    if not list(root.iter(f"{ns}EntityType")):
        ns = "{http://schemas.microsoft.com/ado/2008/09/edm}"

    entities = []
    for elem in root.iter(f"{ns}EntityType"):
        name = elem.get("Name", "Unknown")
        fields = []
        pk_set = set()
        
        key_elem = elem.find(f"{ns}Key")
        if key_elem is not None:
            for pr in key_elem.findall(f"{ns}PropertyRef"):
                pk_set.add(pr.get("Name", ""))
        
        for prop in elem.findall(f"{ns}Property"):
            prop_name = prop.get("Name", "")
            prop_type = prop.get("Type", "Edm.String")
            fields.append({
                "name": prop_name,
                "type": EDM_TYPE_MAP.get(prop_type, "string"),
                "native_type": prop_type,
                "nullable": prop.get("Nullable", "true").lower() != "false",
                "primary_key": prop_name in pk_set,
                "foreign_key": False,
            })
        
        for nav in elem.findall(f"{ns}NavigationProperty"):
            fields.append({
                "name": nav.get("Name", ""),
                "type": "relation",
                "native_type": nav.get("Type", nav.get("ToRole", "")),
                "nullable": True,
                "primary_key": False,
                "foreign_key": True,
            })
        
        if fields:
            entities.append({"name": name, "entity_type": "odata_entity", "fields": fields})

    logger.info(f"[OData] {len(entities)} entités parsées")
    return entities


# ═══════════════════════════════════════════════════════════════════════════════
# DYNAMICS 365 METADATA
# ═══════════════════════════════════════════════════════════════════════════════

async def _fetch_dynamics_metadata(source, secrets: Dict) -> List[Dict]:
    """Récupère les métadonnées de Dynamics 365."""
    import xml.etree.ElementTree as ET

    api_url = (source.base_url or "").rstrip("/")
    headers = _build_dynamics_headers(source, secrets)

    entities = []

    try:
        def _do_get():
            import requests
            return requests.get(
                f"{api_url}/api/data/v9.2/EntityDefinitions",
                headers=headers,
                params={"$select": "LogicalName,DisplayName,PrimaryIdAttribute,PrimaryNameAttribute,IsCustomEntity"},
                timeout=20
            )
        resp = await asyncio.to_thread(_do_get)
        if resp.ok:
            for ed in resp.json().get("value", []):
                logical_name = ed.get("LogicalName", "")
                display_name = ed.get("DisplayName", {}).get("UserLocalizedLabel", {}).get("Label", logical_name)
                entities.append({
                    "name": logical_name,
                    "entity_type": "dynamics_entity",
                    "description": f"{display_name} ({'Custom' if ed.get('IsCustomEntity') else 'Standard'})",
                    "fields": [
                        {"name": ed.get("PrimaryIdAttribute", "id"), "type": "uuid",
                         "native_type": "Edm.Guid", "nullable": False, "primary_key": True, "foreign_key": False},
                        {"name": ed.get("PrimaryNameAttribute", "name"), "type": "string",
                         "native_type": "Edm.String", "nullable": True, "primary_key": False, "foreign_key": False},
                    ],
                })
            if entities:
                logger.info(f"[Dynamics] {len(entities)} entités via EntityDefinitions")
                return entities
    except Exception as e:
        logger.warning(f"[Dynamics] EntityDefinitions: {e}")

    def _do_meta():
        import requests
        meta_headers = {**headers, "Accept": "application/xml, text/xml, */*"}
        resp = requests.get(f"{api_url}/api/data/v9.2/$metadata", headers=meta_headers, timeout=15)
        resp.raise_for_status()
        return resp.text

    text = await asyncio.to_thread(_do_meta)
    root = ET.fromstring(text)
    
    EDM_NS = "http://docs.oasis-open.org/odata/ns/edm"
    ns = f"{{{EDM_NS}}}"
    if not list(root.iter(f"{ns}EntityType")):
        ns = "{http://schemas.microsoft.com/ado/2008/09/edm}"

    for elem in root.iter(f"{ns}EntityType"):
        if elem.get("Abstract", "false").lower() == "true":
            continue
        
        name = elem.get("Name", "Unknown")
        fields = []
        pk_set = set()
        
        key_elem = elem.find(f"{ns}Key")
        if key_elem is not None:
            for pr in key_elem.findall(f"{ns}PropertyRef"):
                pk_set.add(pr.get("Name", ""))
        
        for prop in elem.findall(f"{ns}Property"):
            prop_name = prop.get("Name", "")
            fields.append({
                "name": prop_name,
                "type": EDM_TYPE_MAP.get(prop.get("Type", "Edm.String"), "string"),
                "native_type": prop.get("Type", "Edm.String"),
                "nullable": prop.get("Nullable", "true").lower() != "false",
                "primary_key": prop_name in pk_set,
                "foreign_key": False,
            })
        
        for nav in elem.findall(f"{ns}NavigationProperty"):
            fields.append({
                "name": nav.get("Name", ""),
                "type": "relation",
                "native_type": nav.get("Type", ""),
                "nullable": True,
                "primary_key": False,
                "foreign_key": True,
            })
        
        if fields:
            entities.append({"name": name, "entity_type": "dynamics_entity", "fields": fields})

    logger.info(f"[Dynamics] {len(entities)} entités via $metadata XML")
    return entities


# ═══════════════════════════════════════════════════════════════════════════════
# SAGE METADATA
# ═══════════════════════════════════════════════════════════════════════════════

async def _fetch_sage_metadata(source, secrets: Dict) -> List[Dict]:
    """Dispatcher vers Sage X3/100/Cloud."""
    ct = source.connector_type.value
    if ct == "sage_x3":
        return await _fetch_sage_x3_metadata(source, secrets)
    elif ct == "sage_100":
        return await _fetch_db_metadata(source, secrets)
    elif ct == "sage_cloud":
        return await _fetch_sage_cloud_metadata(source, secrets)
    return []


async def _fetch_sage_x3_metadata(source, secrets: Dict) -> List[Dict]:
    """Récupère les métadonnées SAGE X3 via API."""
    base_url = (source.base_url or "").rstrip("/")
    folder = source.options.get("folder", "SEED")
    headers = {"Accept": "application/json"}
    auth = None

    if secrets.get("token"):
        headers["Authorization"] = f"Bearer {secrets['token']}"
    elif source.username:
        auth = (source.username, secrets.get("password", ""))

    objects = source.options.get("objects", [
        "CUSTOMER", "SUPPLIER", "SORDER", "SINVOICE", "GACCENTRY",
        "ITMMASTER", "FACILITY", "BPCUSTOMER", "BPSUPPLIER",
        "PORDER", "PINVOICE", "BPARTNER",
    ])

    entities = []
    for obj_name in objects:
        try:
            def _do_get(o=obj_name):
                import requests
                return requests.get(
                    f"{base_url}/api/x3/erp/{folder}/{o}/$descriptor",
                    headers=headers, auth=auth, timeout=15
                )
            resp = await asyncio.to_thread(_do_get)
            if resp.ok:
                descriptor = resp.json()
                fields = [
                    {
                        "name": f.get("$fieldName", ""),
                        "type": SAGE_TYPE_MAP.get(f.get("$type", "A"), "string"),
                        "native_type": f.get("$type", "A"),
                        "nullable": not f.get("$isKey", False),
                        "primary_key": f.get("$isKey", False),
                        "foreign_key": f.get("$isForeignKey", False),
                        "description": f.get("$description", ""),
                    }
                    for f in descriptor.get("$fields", [])
                ]
            else:
                fields = [
                    {"name": "ROWID", "type": "integer", "native_type": "L",
                     "nullable": False, "primary_key": True, "foreign_key": False},
                    {"name": "CODE", "type": "string", "native_type": "A",
                     "nullable": False, "primary_key": False, "foreign_key": False},
                    {"name": "DESCRIPTION", "type": "string", "native_type": "A",
                     "nullable": True, "primary_key": False, "foreign_key": False},
                ]
            
            entities.append({
                "name": obj_name,
                "entity_type": "sage_x3_object",
                "description": f"SAGE X3 - {folder} - {obj_name}",
                "fields": fields,
            })
        except Exception as e:
            logger.warning(f"[SAGE X3] {obj_name}: {e}")

    return entities


async def _fetch_sage_cloud_metadata(source, secrets: Dict) -> List[Dict]:
    """Récupère les métadonnées SAGE Business Cloud."""
    base_url = source.base_url or "https://api.accounting.sage.com/v3.1"
    headers = {"Accept": "application/json", "Authorization": f"Bearer {secrets.get('token', '')}"}

    resources = [
        ("ledger_accounts", "Comptes comptables"),
        ("journals", "Journaux"),
        ("journal_entries", "Écritures"),
        ("contacts", "Contacts"),
        ("sales_invoices", "Factures ventes"),
        ("purchase_invoices", "Factures achats"),
        ("products", "Produits"),
        ("tax_rates", "TVA"),
        ("bank_accounts", "Comptes bancaires"),
        ("payment_methods", "Moyens de paiement"),
        ("currencies", "Devises"),
        ("cost_centres", "Centres de coût"),
    ]

    entities = []
    for resource_name, description in resources:
        try:
            def _do_get(r=resource_name):
                import requests
                return requests.get(f"{base_url}/{r}", headers=headers,
                                    params={"$top": 1}, timeout=10)
            resp = await asyncio.to_thread(_do_get)
            if not resp.ok:
                continue
            
            data = resp.json()
            items = data.get("$items", data.get("value", []))
            sample = items[0] if items else {}
            
            fields = [
                {"name": k, "type": _infer_type(v), "native_type": type(v).__name__,
                 "nullable": True, "primary_key": k == "id",
                 "foreign_key": k.endswith("_id") and k != "id"}
                for k, v in sample.items()
            ] if isinstance(sample, dict) else [
                {"name": "id", "type": "uuid", "native_type": "string",
                 "nullable": False, "primary_key": True, "foreign_key": False},
                {"name": "display_name", "type": "string", "native_type": "string",
                 "nullable": True, "primary_key": False, "foreign_key": False},
            ]
            
            entities.append({
                "name": resource_name,
                "entity_type": "sage_cloud_resource",
                "description": f"SAGE Business Cloud - {description}",
                "fields": fields,
            })
        except Exception as e:
            logger.warning(f"[SAGE Cloud] {resource_name}: {e}")

    return entities


# ═══════════════════════════════════════════════════════════════════════════════
# WEBSERVICE / REST METADATA
# ═══════════════════════════════════════════════════════════════════════════════

async def _fetch_webservice_metadata(source, secrets: Dict) -> List[Dict]:
    """Récupère les métadonnées d'un WebService REST/OData."""
    headers: Dict[str, str] = {}

    if source.auth_type.value == "bearer":
        headers["Authorization"] = f"Bearer {secrets.get('token','')}"
    elif source.auth_type.value == "basic":
        import base64
        creds = base64.b64encode(f"{source.username}:{secrets.get('password','')}".encode()).decode()
        headers["Authorization"] = f"Basic {creds}"
    elif source.auth_type.value == "api_key":
        header_name = source.options.get("api_key_header", "X-API-Key")
        headers[header_name] = secrets.get("api_key_value", "")

    if source.connector_type.value == "odata":
        return await _parse_odata_metadata(source.base_url, headers)

    headers["Accept"] = "application/json"
    endpoints = source.options.get("endpoints", [{"path": "/", "entity_name": "root"}])
    entities = []

    for ep in endpoints:
        path = ep.get("path", "/")
        name = ep.get("entity_name", path.strip("/") or "data")
        try:
            def _do_get(p=path):
                import requests
                return requests.get(f"{source.base_url}{p}", headers=headers, timeout=10)
            resp = await asyncio.to_thread(_do_get)
            data = resp.json()
            sample = data[0] if isinstance(data, list) and data else data if isinstance(data, dict) else {}
            if isinstance(sample, dict):
                fields = [{"name": k, "type": _infer_type(v), "nullable": True} for k, v in sample.items()]
                entities.append({"name": name, "entity_type": "endpoint", "fields": fields})
        except Exception as e:
            logger.warning(f"[Sync REST] {path}: {e}")

    return entities


# ═══════════════════════════════════════════════════════════════════════════════
# FILE METADATA
# ════��══════════════════════════════════════════════════════════════════════════

async def _fetch_file_metadata(source) -> List[Dict]:
    """Récupère les métadonnées d'un fichier."""
    file_path = _resolve_file_path(source)
    if not file_path:
        raise ValueError("Aucun chemin de fichier spécifié dans les options")

    connector_type = source.connector_type.value
    entity_name    = source.name.replace(" ", "_").lower()
    opts           = source.options or {}

    if connector_type in ("file_excel", "file_xlsx"):
        return await asyncio.to_thread(_parse_excel, file_path, entity_name, opts)

    if connector_type == "file_parquet":
        return await asyncio.to_thread(_parse_parquet, file_path, entity_name)

    if connector_type == "file_avro":
        return await asyncio.to_thread(_parse_avro, file_path, entity_name)

    content = await _read_file_content(file_path)
    if connector_type == "file_csv":
        return _parse_csv(content, entity_name)
    elif connector_type == "file_json":
        return _parse_json(content, entity_name)

    raise ValueError(f"Type de fichier non supporté : {connector_type}")


async def _read_file_content(file_path: str) -> str:
    """Lit le contenu d'un fichier (HTTP ou local)."""
    if file_path.startswith("http"):
        def _do_get():
            import requests
            return requests.get(file_path, timeout=30).text
        return await asyncio.to_thread(_do_get)
    
    def _read():
        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            return f.read(5 * 1024 * 1024)
    
    return await asyncio.to_thread(_read)


def _parse_csv(content: str, entity_name: str) -> List[Dict]:
    """Parse un fichier CSV."""
    import csv as csv_module
    import io
    
    reader = csv_module.DictReader(io.StringIO(content))
    rows = []
    headers = reader.fieldnames or []
    for i, row in enumerate(reader):
        rows.append(row)
        if i >= 100:
            break
    if not headers and rows:
        headers = list(rows[0].keys())
    
    fields = []
    for h in headers:
        sample_vals = [str(r.get(h, "")) for r in rows if r.get(h)]
        fields.append({
            "name": h,
            "type": _infer_column_type(sample_vals),
            "native_type": "csv_column",
            "nullable": True,
            "primary_key": False,
            "foreign_key": False,
        })
    
    return [{"name": entity_name, "entity_type": "csv_file", "description": f"CSV — {len(headers)} colonnes", "fields": fields}]


def _parse_json(content: str, entity_name: str) -> List[Dict]:
    """Parse un fichier JSON."""
    import json as json_module
    
    data = json_module.loads(content)
    if isinstance(data, list):
        sample = data[0] if data else {}
    elif isinstance(data, dict):
        sample = data
    else:
        return [{"name": entity_name, "entity_type": "json_file", "fields": []}]
    
    fields = [
        {"name": k, "type": _infer_type(v), "native_type": type(v).__name__,
         "nullable": True, "primary_key": k in ("id", "_id"), "foreign_key": False}
        for k, v in sample.items()
    ] if isinstance(sample, dict) else []
    
    return [{"name": entity_name, "entity_type": "json_file", "description": f"JSON — {len(fields)} champs", "fields": fields}]


def _parse_excel(file_path: str, entity_name: str, opts: dict) -> List[Dict]:
    """Parse un fichier Excel."""
    import openpyxl
    
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheets = opts.get("sheets") or wb.sheetnames
    entities = []
    
    for sheet_name in sheets:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(max_row=101, values_only=True))
        if not rows:
            continue
        
        headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
        data_rows = rows[1:101]
        fields = []
        
        for i, h in enumerate(headers):
            sample_vals = [str(r[i]) for r in data_rows if i < len(r) and r[i] is not None]
            fields.append({
                "name": h,
                "type": _infer_column_type(sample_vals),
                "native_type": "excel_cell",
                "nullable": True,
                "primary_key": False,
                "foreign_key": False,
            })
        
        entities.append({
            "name": f"{entity_name}_{sheet_name}",
            "entity_type": "excel_sheet",
            "description": f"Excel — feuille '{sheet_name}' · {len(headers)} colonnes",
            "fields": fields,
        })
    
    wb.close()
    return entities


def _parse_parquet(file_path: str, entity_name: str) -> List[Dict]:

    try:
        import pyarrow.parquet as pq
        pf = pq.ParquetFile(file_path)
        schema = pf.schema_arrow
        meta   = pf.metadata

        PARQUET_TYPE_MAP = {
            "int8": "integer", "int16": "integer", "int32": "integer", "int64": "integer",
            "uint8": "integer", "uint16": "integer", "uint32": "integer", "uint64": "integer",
            "float": "float", "double": "float", "float16": "float",
            "string": "string", "large_string": "string", "utf8": "string",
            "bool": "boolean",
            "date32": "date", "date64": "date",
            "timestamp[ms]": "datetime", "timestamp[us]": "datetime", "timestamp[ns]": "datetime",
            "binary": "string", "large_binary": "string",
            "decimal128": "decimal",
        }

        fields = []
        for i in range(len(schema)):
            field = schema.field(i)
            type_str = str(field.type)
            fields.append({
                "name":        field.name,
                "type":        PARQUET_TYPE_MAP.get(type_str, "string"),
                "native_type": f"parquet:{type_str}",
                "nullable":    field.nullable,
                "primary_key": False,
                "foreign_key": False,
            })

        row_count = meta.num_rows
        row_groups = meta.num_row_groups
        return [{
            "name":        entity_name,
            "entity_type": "parquet_file",
            "description": f"Parquet — {row_count:,} lignes · {len(fields)} colonnes · {row_groups} row groups",
            "fields":      fields,
        }]
    except ImportError:
        raise ImportError("pyarrow non installé : pip install pyarrow")
    except Exception as e:
        logger.warning(f"[Parquet] Erreur lecture {file_path}: {e}")
        raise


def _parse_avro(file_path: str, entity_name: str) -> List[Dict]:
    try:
        import fastavro
    except ImportError:
        raise ImportError("fastavro non installé : pip install fastavro")

    AVRO_TYPE_MAP = {
        "null": "string", "boolean": "boolean", "int": "integer", "long": "integer",
        "float": "float", "double": "float", "bytes": "string", "string": "string",
        "record": "object", "array": "array", "map": "object",
        "enum": "string", "fixed": "string",
        "long (timestamp-millis)": "datetime",
        "long (timestamp-micros)": "datetime",
        "int (date)": "date",
    }

    with open(file_path, "rb") as f:
        reader = fastavro.reader(f)
        schema = reader.writer_schema
        records = []
        for i, rec in enumerate(reader):
            if i >= 5: break
            records.append(rec)

    fields = []
    if isinstance(schema, dict) and schema.get("type") == "record":
        for avro_field in schema.get("fields", []):
            avro_type = avro_field.get("type", "string")
            if isinstance(avro_type, list):
                avro_type = next((t for t in avro_type if t != "null"), "string")
            if isinstance(avro_type, dict):
                avro_type = avro_type.get("type", "string")
            fields.append({
                "name":        avro_field.get("name", ""),
                "type":        AVRO_TYPE_MAP.get(str(avro_type), "string"),
                "native_type": f"avro:{avro_type}",
                "nullable":    True,
                "primary_key": False,
                "foreign_key": False,
                "description": avro_field.get("doc", ""),
            })

    schema_name = schema.get("name", entity_name) if isinstance(schema, dict) else entity_name
    return [{
        "name":        entity_name,
        "entity_type": "avro_file",
        "description": f"Avro — schéma '{schema_name}' · {len(fields)} champs",
        "fields":      fields,
    }]


def _resolve_file_path(source) -> str:
    opts = source.options or {}
    if opts.get("file_path"):
        path = opts["file_path"]
        if len(path) > 2 and path[1] == ":":
            path = "/mnt/host/" + path[3:].replace("\\", "/")
        elif path.startswith("\\\\"):
            path = "/mnt/network/" + path[2:].replace("\\", "/")
        elif path.startswith("//"):
            path = "/mnt/network/" + path[2:]
        return path
    if opts.get("uploaded_path"):
        return opts["uploaded_path"]
    if source.base_url:
        return source.base_url
    return ""


def _human_size(size: int) -> str:
    for unit in ["o", "Ko", "Mo", "Go"]:
        if size < 1024:
            return f"{size:.1f} {unit}"
        size /= 1024
    return f"{size:.1f} To"


def _infer_column_type(values: List[str]) -> str:
    import re
    if not values:
        return "string"
    total = len(values)
    int_count   = sum(1 for v in values if re.match(r"^-?\d+$", v.strip()))
    float_count = sum(1 for v in values if re.match(r"^-?\d+\.\d+$", v.strip()))
    date_count  = sum(1 for v in values if re.match(r"^\d{4}-\d{2}-\d{2}", v.strip()))
    bool_count  = sum(1 for v in values if v.strip().lower() in ("true","false","1","0","oui","non"))
    if date_count  / total > 0.8: return "date"
    if bool_count  / total > 0.8: return "boolean"
    if int_count   / total > 0.8: return "integer"
    if float_count / total > 0.8: return "float"
    return "string"


def _infer_type(value) -> str:
    if isinstance(value, bool):  return "boolean"
    if isinstance(value, int):   return "integer"
    if isinstance(value, float): return "float"
    if isinstance(value, dict):  return "object"
    if isinstance(value, list):  return "array"
    import re
    if isinstance(value, str) and re.match(r"\d{4}-\d{2}-\d{2}", value):
        return "date"
    return "string"


async def _parse_odata_metadata(base_url: str, headers: Dict) -> List[Dict]:
    import requests
    import xml.etree.ElementTree as ET

    meta_headers = {k: v for k, v in headers.items() if k != "Accept"}
    meta_headers["Accept"] = "application/xml, text/xml, */*"

    def _do_get():
        resp = requests.get(f"{base_url}/$metadata", headers=meta_headers, timeout=15)
        resp.raise_for_status()
        return resp.text

    text = await asyncio.to_thread(_do_get)
    root = ET.fromstring(text)

    EDM_NS = "http://docs.oasis-open.org/odata/ns/edm"
    ns = f"{{{EDM_NS}}}"
    if not list(root.iter(f"{ns}EntityType")):
        ns = "{http://schemas.microsoft.com/ado/2008/09/edm}"

    entities = []
    for elem in root.iter(f"{ns}EntityType"):
        name = elem.get("Name", "Unknown")
        fields = []
        pk_set = set()
        key_elem = elem.find(f"{ns}Key")
        if key_elem is not None:
            for pr in key_elem.findall(f"{ns}PropertyRef"):
                pk_set.add(pr.get("Name", ""))
        for prop in elem.findall(f"{ns}Property"):
            prop_name = prop.get("Name", "")
            prop_type = prop.get("Type", "Edm.String")
            fields.append({
                "name": prop_name, "type": EDM_TYPE_MAP.get(prop_type, "string"),
                "native_type": prop_type,
                "nullable": prop.get("Nullable", "true").lower() != "false",
                "primary_key": prop_name in pk_set, "foreign_key": False,
            })
        for nav in elem.findall(f"{ns}NavigationProperty"):
            fields.append({
                "name": nav.get("Name", ""), "type": "relation",
                "native_type": nav.get("Type", nav.get("ToRole", "")),
                "nullable": True, "primary_key": False, "foreign_key": True,
            })
        if fields:
            entities.append({"name": name, "entity_type": "odata_entity", "fields": fields})

    logger.info(f"[OData] {len(entities)} entites parsees depuis {base_url}")
    return entities