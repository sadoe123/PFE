"""
OnePilot – Metadata Enricher §2.2.1 completion
Sauvegarde les métadonnées enrichies manquantes :
- Commentaires/descriptions COMMENT ON
- Dépendances views → tables
- Statistiques cardinalité dans le catalogue
- Formules Excel
"""
from __future__ import annotations

import json
import logging
from typing import Dict, List, Optional
from uuid import UUID

import asyncpg

logger = logging.getLogger(__name__)


class MetadataEnricher:
    """
    Enrichit les entités déjà indexées avec les métadonnées
    qui n'étaient pas sauvegardées lors du sync initial.
    """

    def __init__(self, pg_pool: asyncpg.Pool):
        self.pg = pg_pool

    # ── 1. Sauvegarder commentaires/descriptions ──────────────────────

    async def save_descriptions(
        self, source_id: UUID, descriptions: Dict[str, Dict[str, Optional[str]]]
    ) -> int:
        """
        Sauvegarde les descriptions extraites via COMMENT ON.
        descriptions = {
            "table_name": {
                "_table": "Description de la table",
                "col_name": "Description de la colonne"
            }
        }
        """
        count = 0
        for table_name, desc_map in descriptions.items():
            # Description table
            table_desc = desc_map.get("_table")
            if table_desc:
                result = await self.pg.execute("""
                    UPDATE source_entities
                    SET description = $1
                    WHERE source_id = $2
                      AND name = $3
                      AND (description IS NULL OR description = '')
                """, table_desc, source_id, table_name)
                if result != "UPDATE 0":
                    count += 1

            # Descriptions colonnes
            for col_name, col_desc in desc_map.items():
                if col_name == "_table" or not col_desc:
                    continue
                await self.pg.execute("""
                    UPDATE entity_fields ef
                    SET display_name = $1
                    FROM source_entities se
                    WHERE ef.entity_id = se.id
                      AND se.source_id = $2
                      AND se.name = $3
                      AND ef.name = $4
                      AND (ef.display_name IS NULL OR ef.display_name = ef.name)
                """, col_desc[:255], source_id, table_name, col_name)

        logger.info(f"[MetadataEnricher] {count} descriptions sauvegardées pour source {source_id}")
        return count

    # ── 2. Sauvegarder dépendances views → tables ─────────────────────

    async def save_dependencies(
        self, source_id: UUID, dep_map: Dict[str, List[str]]
    ) -> int:
        """
        Sauvegarde les dépendances entre objets (view → table).
        dep_map = {"view_name": ["table1", "table2"]}
        """
        count = 0
        for obj_name, deps in dep_map.items():
            if not deps:
                continue

            # Met à jour le metadata JSON de l'entité
            entity = await self.pg.fetchrow("""
                SELECT id, metadata FROM source_entities
                WHERE source_id = $1 AND name = $2
            """, source_id, obj_name)

            if not entity:
                continue

            meta = {}
            try:
                raw = entity["metadata"]
                meta = json.loads(raw) if isinstance(raw, str) else (raw or {})
            except Exception:
                pass

            meta["dependencies"] = deps
            meta["dependency_count"] = len(deps)

            await self.pg.execute("""
                UPDATE source_entities
                SET metadata = $1
                WHERE id = $2
            """, json.dumps(meta), entity["id"])
            count += 1

        logger.info(f"[MetadataEnricher] {count} dépendances sauvegardées pour source {source_id}")
        return count

    # ── 3. Sauvegarder statistiques cardinalité ───────────────────────

    async def save_cardinality_stats(
        self, source_id: UUID, stats_map: Dict[str, Dict]
    ) -> int:
        """
        Sauvegarde les statistiques de cardinalité dans les entités.
        stats_map = {
            "table_name": {
                "row_count": 12345,
                "size_kb": 256,
                "col_stats": {
                    "col_name": {
                        "distinct_count": 100,
                        "null_count": 5,
                        "null_pct": 0.04
                    }
                }
            }
        }
        """
        count = 0
        for table_name, stats in stats_map.items():
            row_count = stats.get("row_count")
            size_kb = stats.get("size_kb")
            col_stats = stats.get("col_stats", {})

            # Met à jour row_count sur l'entité
            updates = []
            if row_count is not None:
                updates.append(("row_count", row_count))

            entity = await self.pg.fetchrow("""
                SELECT id, metadata FROM source_entities
                WHERE source_id = $1 AND name = $2
            """, source_id, table_name)

            if not entity:
                continue

            # Met à jour row_count
            if row_count is not None:
                await self.pg.execute("""
                    UPDATE source_entities SET row_count = $1 WHERE id = $2
                """, row_count, entity["id"])

            # Met à jour metadata avec size_kb et col_stats
            if size_kb or col_stats:
                meta = {}
                try:
                    raw = entity["metadata"]
                    meta = json.loads(raw) if isinstance(raw, str) else (raw or {})
                except Exception:
                    pass

                if size_kb:
                    meta["size_kb"] = size_kb
                if col_stats:
                    meta["col_stats"] = col_stats

                await self.pg.execute("""
                    UPDATE source_entities SET metadata = $1 WHERE id = $2
                """, json.dumps(meta), entity["id"])

            count += 1

        logger.info(f"[MetadataEnricher] {count} stats cardinalité sauvegardées pour source {source_id}")
        return count

    # ── 4. Extraire et sauvegarder formules Excel ─────────────────────

    async def save_excel_formulas(
        self, source_id: UUID, formulas: Dict[str, List[Dict]]
    ) -> int:
        """
        Sauvegarde les formules Excel détectées.
        formulas = {
            "sheet_name": [
                {"cell": "A1", "formula": "=SUM(B1:B10)", "column": "Total"}
            ]
        }
        """
        count = 0
        for sheet_name, sheet_formulas in formulas.items():
            if not sheet_formulas:
                continue

            entity = await self.pg.fetchrow("""
                SELECT id, metadata FROM source_entities
                WHERE source_id = $1 AND name = $2
            """, source_id, sheet_name)

            if not entity:
                continue

            meta = {}
            try:
                raw = entity["metadata"]
                meta = json.loads(raw) if isinstance(raw, str) else (raw or {})
            except Exception:
                pass

            meta["excel_formulas"] = sheet_formulas[:50]  # cap sécurité
            meta["formula_count"] = len(sheet_formulas)

            await self.pg.execute("""
                UPDATE source_entities SET metadata = $1 WHERE id = $2
            """, json.dumps(meta), entity["id"])
            count += 1

        logger.info(f"[MetadataEnricher] {count} formules Excel sauvegardées pour source {source_id}")
        return count

    # ── 5. Extraction formules depuis fichier Excel ────────────────────

    @staticmethod
    def extract_excel_formulas(filepath: str) -> Dict[str, List[Dict]]:
        """
        Extrait les formules d'un fichier Excel.
        Retourne dict par feuille avec liste des formules trouvées.
        """
        try:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, data_only=False)
            result = {}
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                formulas = []
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                            formulas.append({
                                "cell":    cell.coordinate,
                                "formula": cell.value[:200],
                                "column":  ws.cell(row=1, column=cell.column).value,
                            })
                            if len(formulas) >= 100:
                                break
                    if len(formulas) >= 100:
                        break
                if formulas:
                    result[sheet_name] = formulas
            wb.close()
            return result
        except Exception as e:
            logger.warning(f"[MetadataEnricher] Excel formulas error: {e}")
            return {}

    # ── 6. Extraction cardinalité PostgreSQL ──────────────────────────

    @staticmethod
    def extract_pg_cardinality_sql() -> str:
        """Requête SQL pour extraire les stats de cardinalité PostgreSQL."""
        return """
            SELECT
                c.relname                           AS table_name,
                c.reltuples::bigint                 AS row_count,
                pg_total_relation_size(c.oid)/1024  AS size_kb,
                s.attname                           AS col_name,
                s.n_distinct,
                s.null_frac
            FROM pg_class c
            JOIN pg_namespace n ON n.oid = c.relnamespace
            LEFT JOIN pg_stats s ON s.tablename = c.relname
                                AND s.schemaname = n.nspname
            WHERE n.nspname = :schema
              AND c.relkind = 'r'
            ORDER BY c.relname, s.attname
        """

    # ── 7. Extraction cardinalité MSSQL ───────────────────────────────

    @staticmethod
    def extract_mssql_cardinality_sql() -> str:
        """Requête SQL pour extraire les stats de cardinalité SQL Server."""
        return """
            SELECT
                t.name                          AS table_name,
                p.rows                          AS row_count,
                SUM(a.total_pages) * 8          AS size_kb
            FROM sys.tables t
            JOIN sys.indexes i ON i.object_id = t.object_id AND i.index_id <= 1
            JOIN sys.partitions p ON p.object_id = t.object_id AND p.index_id = i.index_id
            JOIN sys.allocation_units a ON a.container_id = p.partition_id
            WHERE SCHEMA_NAME(t.schema_id) = :schema
            GROUP BY t.name, p.rows
            ORDER BY t.name
        """